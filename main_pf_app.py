#!/usr/bin/env python3
"""
Main Portfolio Analysis Application
Lightweight modular portfolio analysis system with HTML reporting

IMPORTANT: DATA SEPARATION PRINCIPLE
-------------------------------------
This module does NOT fetch market data. It only:
  1. Loads cached data from price_cache/
  2. Performs calculations and analysis
  3. Generates reports and charts

To update market data, use: python3 update_portfolio_data.py

SYSTEM STANDARD: Date Column Naming Convention
-----------------------------------------------
All historical data uses lowercase 'date' column (not 'Date').
This is the natural output of pandas reset_index() on DatetimeIndex.
All modules (pf_drag_analyzer, pf_optimizer, interactive_filter) 
automatically convert 'Date' to 'date' for consistency.
"""

import pandas as pd
import numpy as np
import os
import gc
import glob
from datetime import datetime, date
from typing import Dict, List, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed

# Import our modular components
from data_fetcher import DataManager
from pf_manager import PortfolioManager
from technical_indicators import TechnicalAnalyzer
from html_report_generator import HTMLReportGenerator
from interactive_filter import InteractiveFilter
from stock_scorer import StockScorer
from signal_engine import SignalEngine
from minervini_analyzer import MinerviniAnalyzer


def _default_portfolio_file() -> str:
    """Get portfolio filename from config.json."""
    from config_manager import get_config
    return get_config().get_setting('system_settings.portfolio_file', 'dpsr_report.xls.xlsx')


class PortfolioApp:
    """Main portfolio analysis application"""
    
    def __init__(self, portfolio_file: str = None):
        if portfolio_file is None:
            portfolio_file = _default_portfolio_file()
        print("🚀 Initializing Portfolio Analysis System...")
        print("=" * 60)
        
        # Initialize components
        self.data_manager = DataManager()
        self.portfolio_manager = PortfolioManager(self.data_manager, portfolio_file)
        self.technical_analyzer = TechnicalAnalyzer()
        self.html_generator = HTMLReportGenerator()
        
        # State variables
        self.analysis_results = pd.DataFrame()
        self.insights = {}
        self.charts = {}
        self.comprehensive_dataset = pd.DataFrame()
        self.interactive_filter = None  # Will be initialized after dataset generation
        
        print("✅ System initialized successfully!")
        print("📁 Data cache directory:", self.data_manager.cache_dir)
        print("📊 Portfolio file:", portfolio_file)
        print("=" * 60)
    
    def _load_cached_historical_data(self, force_reload: bool = False) -> bool:
        """Load historical data from cache ONLY (no fetching)"""
        
        # Check if already loaded and not forcing reload
        if not force_reload and hasattr(self, 'historical_data') and self.historical_data is not None and not self.historical_data.empty:
            print("✅ Using already loaded historical data")
            return True
        
        # Ensure portfolio data is loaded
        if not hasattr(self, 'portfolio_data') or self.portfolio_data is None or self.portfolio_data.empty:
            print("⚠️  Portfolio data not loaded, loading now...")
            if not hasattr(self.portfolio_manager, 'portfolio_data') or self.portfolio_manager.portfolio_data.empty:
                return False
            self.portfolio_data = self.portfolio_manager.portfolio_data
        
        # Load from cache only (no network calls)
        print("📁 Loading historical data from cache...")
        portfolio_symbols = self.portfolio_data['Symbol'].tolist()
        self.historical_data = self.portfolio_manager.data_manager.get_multiple_stocks_data(portfolio_symbols)
        
        if not self.historical_data.empty:
            print(f"✅ Loaded {len(self.historical_data)} records for {len(portfolio_symbols)} stocks from cache")
            return True
        else:
            print("❌ No cached data found. Run: python3 update_portfolio_data.py")
            return False
    
    def load_portfolio(self, filepath: str = None) -> bool:
        """Load portfolio data from Excel file"""
        print("\n📊 Loading portfolio data...")
        
        portfolio_data = self.portfolio_manager.load_portfolio_from_excel(filepath)
        
        if not portfolio_data.empty:
            print(f"✅ Portfolio loaded: {len(portfolio_data)} stocks")
            return True
        else:
            print("❌ Failed to load portfolio data")
            return False
    
    def check_data_availability(self) -> bool:
        """Check if cached data is available for analysis.
        Returns True even if a few stocks are missing (they'll be skipped).
        Returns False only if the majority of stocks lack data.
        """
        print("\n🔍 Checking data availability...")
        
        # Get portfolio symbols
        if not hasattr(self.portfolio_manager, 'portfolio_data') or self.portfolio_manager.portfolio_data.empty:
            print("❌ No portfolio data loaded")
            return False
        
        symbols = self.portfolio_manager.portfolio_data['Symbol'].tolist()
        
        # Check cache directory
        import glob
        cache_files = glob.glob('price_cache/*_data.pkl')
        cached_symbols = set([os.path.basename(f).replace('_data.pkl', '') for f in cache_files])
        
        missing_symbols = []
        for symbol in symbols:
            symbol_ns = f"{symbol}.NS" if not symbol.endswith(('.NS', '.BO')) else symbol
            symbol_bo = symbol_ns.replace('.NS', '.BO')
            
            if symbol_ns not in cached_symbols and symbol_bo not in cached_symbols:
                missing_symbols.append(symbol)
        
        if missing_symbols:
            pct_missing = len(missing_symbols) / len(symbols) * 100
            if pct_missing > 50:
                print(f"❌ {len(missing_symbols)}/{len(symbols)} ({pct_missing:.0f}%) symbols missing from cache — too many to proceed")
                print("💡 Run: python3 update_portfolio_data.py")
                if len(missing_symbols) <= 20:
                    print(f"   Missing: {', '.join(missing_symbols)}")
                return False
            else:
                print(f"⚠️  {len(missing_symbols)} symbol(s) not in cache (will be skipped in reports):")
                for s in missing_symbols:
                    print(f"   ⏭️  {s}")
                print(f"💡 To fetch data for new stocks, run: python3 update_portfolio_data.py")
                print(f"✅ Proceeding with {len(symbols) - len(missing_symbols)}/{len(symbols)} stocks that have cached data")
                return True
        else:
            print(f"✅ All {len(symbols)} symbols have cached data")
            return True
    
    def run_analysis(self) -> bool:
        """Run comprehensive portfolio analysis using master dataset"""
        print("\n🔍 Running comprehensive portfolio analysis...")
        
        try:
            # Ensure we have portfolio data
            if not hasattr(self, 'portfolio_data') or self.portfolio_data is None:
                self.portfolio_data = self.portfolio_manager.portfolio_data
                
            # Load historical data from cache
            if not self._load_cached_historical_data():
                print("❌ Failed to load cached data")
                return False
            
            # Generate comprehensive master dataset (uses cached historical_data)
            self.generate_comprehensive_dataset()
            
            if self.comprehensive_dataset.empty:
                print("❌ No data available for analysis")
                return False
            
            # Use comprehensive dataset as source for all analysis (reference, not copy)
            self.analysis_results = self.comprehensive_dataset
            
            # Generate insights from comprehensive dataset
            self.generate_insights()
            
            print("✅ Portfolio analysis complete for {} stocks".format(len(self.analysis_results)))
            return True
            
        except Exception as e:
            print(f"❌ Error in portfolio analysis: {e}")
            return False
    
    def _calculate_portfolio_metrics(self, portfolio_row: pd.Series, tech_analysis: dict) -> dict:
        """Calculate portfolio-specific metrics"""
        metrics = {}
        
        try:
            # Basic portfolio data
            dp_bal = portfolio_row.get('DP_Bal', 0)
            hold_price = portfolio_row.get('Hold_Price', 0)
            allocation = portfolio_row.get('Percentage_Allocation', 0)
            
            # Current market values
            current_price = tech_analysis.get('current_price', 0)
            
            # Financial calculations
            buy_value = dp_bal * hold_price
            current_value = dp_bal * current_price
            profit_loss = current_value - buy_value
            profit_loss_pct = ((current_value - buy_value) / buy_value) * 100 if buy_value > 0 else 0
            
            metrics.update({
                'DP_Bal': dp_bal,
                'Hold_Price': hold_price,
                'Buy_Value': buy_value,
                'Current_Value': current_value,
                'Profit_Loss': profit_loss,
                'Profit_Loss_Pct': profit_loss_pct,
                'Allocation_Pct': allocation
            })
            
        except Exception as e:
            print(f"      ⚠️  Error calculating portfolio metrics: {e}")
        
        return metrics
    
    def generate_insights(self) -> bool:
        """Generate portfolio insights from comprehensive dataset"""
        print("\n💡 Generating portfolio insights...")
        
        try:
            # Check data freshness before generating insights
            self.check_data_freshness_before_operation("insight generation")
            
            # Ensure comprehensive dataset is available
            if not hasattr(self, 'comprehensive_dataset') or self.comprehensive_dataset.empty:
                print("🔄 Generating comprehensive dataset for insights...")
                self.generate_comprehensive_dataset()
            
            # Use comprehensive dataset for insights
            dataset = self.comprehensive_dataset
            
            if dataset.empty:
                print("❌ No data available for insights")
                return False
            
            # Portfolio summary from comprehensive dataset
            portfolio_summary = {
                'total_stocks': len(dataset),
                'total_investment': (dataset['DP_Bal'] * dataset['Hold_Price']).sum(),
                'current_value': (dataset['DP_Bal'] * dataset['CMP']).sum(),
                'total_pnl': dataset['Profit/Loss'].sum(),
                'portfolio_return_pct': 0
            }
            
            if portfolio_summary['total_investment'] > 0:
                portfolio_summary['portfolio_return_pct'] = (
                    (portfolio_summary['current_value'] - portfolio_summary['total_investment']) / 
                    portfolio_summary['total_investment']
                ) * 100
            
            # Performance analysis from comprehensive dataset
            best_performer = dataset.loc[dataset['Profit/Loss'].idxmax()] if len(dataset) > 0 else {}
            worst_performer = dataset.loc[dataset['Profit/Loss'].idxmin()] if len(dataset) > 0 else {}
            
            performance = {
                'best_performer': {
                    'symbol': best_performer.get('Symbol', 'N/A'),
                    'return_pct': best_performer.get('Profit/Loss', 0) / best_performer.get('Hold_Price', 1) * 100 if 'Hold_Price' in best_performer and best_performer.get('Hold_Price', 0) > 0 else 0
                },
                'worst_performer': {
                    'symbol': worst_performer.get('Symbol', 'N/A'),
                    'return_pct': worst_performer.get('Profit/Loss', 0) / worst_performer.get('Hold_Price', 1) * 100 if 'Hold_Price' in worst_performer and worst_performer.get('Hold_Price', 0) > 0 else 0
                }
            }
            
            # Risk analysis from comprehensive dataset
            avg_rsi = dataset['RSI'].mean() if 'RSI' in dataset.columns else 50
            avg_volatility = dataset['Standard_Deviation'].mean() if 'Standard_Deviation' in dataset.columns else 0
            
            risk_analysis = {
                'avg_rsi': avg_rsi,
                'avg_portfolio_volatility': avg_volatility,
                'high_volatility_stocks': len(dataset[dataset['Standard_Deviation'] > 30]) if 'Standard_Deviation' in dataset.columns else 0,
                'low_volatility_stocks': len(dataset[dataset['Standard_Deviation'] < 10]) if 'Standard_Deviation' in dataset.columns else 0,
                'risk_level': 'HIGH' if avg_volatility > 30 else 'MEDIUM' if avg_volatility > 15 else 'LOW'
            }
            
            # Technical analysis
            technical_analysis = {
                'oversold_rsi': len(dataset[dataset['RSI'] < 30]) if 'RSI' in dataset.columns else 0,
                'overbought_rsi': len(dataset[dataset['RSI'] > 70]) if 'RSI' in dataset.columns else 0,
                'neutral_stocks': len(dataset[(dataset['RSI'] >= 30) & (dataset['RSI'] <= 70)]) if 'RSI' in dataset.columns else 0
            }

            self.insights = {
                'portfolio_summary': portfolio_summary,
                'performance': performance,
                'risk_analysis': risk_analysis,
                'technical_analysis': technical_analysis
            }
            
            print("✅ Insights generated successfully")
            return True
            
        except Exception as e:
            print(f"❌ Error generating insights: {e}")
            return False
    
    def create_visualizations(self) -> dict:
        """Create all portfolio visualizations from comprehensive dataset"""
        print("\n🎨 Creating portfolio visualizations...")
        
        # Use existing comprehensive dataset if available, don't regenerate
        if not hasattr(self, 'comprehensive_dataset') or self.comprehensive_dataset.empty:
            print("❌ No comprehensive dataset available for visualizations")
            return {}
        
        # Use comprehensive dataset for chart generation
        try:
            self.charts = self.html_generator.create_portfolio_charts(
                self.comprehensive_dataset, 
                self.portfolio_manager
            )
            print("✅ All portfolio charts created successfully")
            return self.charts
        except Exception as e:
            print(f"❌ Error creating visualizations: {e}")
            return {}
    
    def generate_html_report(self) -> str:
        """Generate comprehensive HTML report"""
        print("\n📄 Generating HTML portfolio report...")
        
        try:
            # Check data freshness before generating report
            self.check_data_freshness_before_operation("HTML report generation")
            
            # Ensure we have insights
            if not self.insights:
                self.generate_insights()
            
            # Ensure we have charts
            if not self.charts:
                self.create_visualizations()
            
            # Generate HTML report (saves automatically to reports folder)
            html_report = self.html_generator.generate_html_report(
                self.analysis_results,
                self.insights,
                self.charts
            )
            
            # Create a text summary
            text_summary = self._create_text_summary()
            timestamp = date.today().strftime('%Y%m%d')
            text_filename = f"reports/portfolio_summary_{timestamp}.txt"
            
            # Ensure reports directory exists
            os.makedirs('reports', exist_ok=True)
            
            with open(text_filename, 'w', encoding='utf-8') as f:
                f.write(text_summary)
            
            print(f"📄 Text summary saved to {text_filename}")
            
            return html_report
            
        except Exception as e:
            print(f"❌ Error generating report: {e}")
            return ""
    
    def _create_text_summary(self) -> str:
        """Create a text summary report"""
        if not self.insights:
            return "No analysis data available for summary."
        
        lines = []
        lines.append("=" * 60)
        lines.append("📊 PORTFOLIO ANALYSIS SUMMARY")
        lines.append("=" * 60)
        lines.append(f"📅 Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("")
        
        # Portfolio overview
        if 'portfolio_summary' in self.insights:
            ps = self.insights['portfolio_summary']
            lines.append("💼 PORTFOLIO OVERVIEW")
            lines.append("-" * 30)
            lines.append(f"Total Stocks: {ps['total_stocks']}")
            lines.append(f"Total Investment: ₹{ps['total_investment']:,.2f}")
            lines.append(f"Current Value: ₹{ps['current_value']:,.2f}")
            lines.append(f"Total P&L: ₹{ps['total_pnl']:,.2f}")
            lines.append(f"Portfolio Return: {ps['portfolio_return_pct']:.2f}%")
            lines.append("")
        
        # Performance highlights
        if 'performance' in self.insights:
            perf = self.insights['performance']
            lines.append("🏆 PERFORMANCE HIGHLIGHTS")
            lines.append("-" * 30)
            lines.append(f"Best Performer: {perf['best_performer']['symbol']} ({perf['best_performer']['return_pct']:.2f}%)")
            lines.append(f"Worst Performer: {perf['worst_performer']['symbol']} ({perf['worst_performer']['return_pct']:.2f}%)")
            lines.append("")
        
        # Risk analysis
        if 'risk_analysis' in self.insights:
            risk = self.insights['risk_analysis']
            lines.append("⚠️  RISK ANALYSIS")
            lines.append("-" * 30)
            lines.append(f"Average Volatility: {risk['avg_portfolio_volatility']:.2%}")
            lines.append(f"High Volatility Stocks: {risk['high_volatility_stocks']}")
            lines.append(f"Low Volatility Stocks: {risk['low_volatility_stocks']}")
            lines.append("")
        
        # Technical analysis
        if 'technical_analysis' in self.insights:
            tech = self.insights['technical_analysis']
            lines.append("📈 TECHNICAL ANALYSIS")
            lines.append("-" * 30)
            lines.append(f"Oversold Stocks (RSI < 30): {tech['oversold_rsi']}")
            lines.append(f"Overbought Stocks (RSI > 70): {tech['overbought_rsi']}")
            lines.append(f"Neutral Stocks: {tech['neutral_stocks']}")
            lines.append("")
        
        lines.append("=" * 60)
        lines.append("Generated by Portfolio Analysis System")
        lines.append("=" * 60)
        
        return "\n".join(lines)
    
    def run_complete_analysis(self, portfolio_file: str = None) -> bool:
        """Run complete end-to-end analysis workflow (using cached data only)
        
        Note: This method does NOT fetch market data. To update data first, run:
              python3 update_portfolio_data.py
        """
        print("🚀 Starting complete portfolio analysis workflow...")
        print("💡 Using cached data only (no network calls)")
        print("=" * 80)
        
        success = True
        
        try:
            # Step 0: Auto-clean old reports (keep only today's)
            print("\n🧹 STEP 0: Cleaning Old Reports")
            self._cleanup_old_reports()
            
            # Step 1: Load portfolio
            print("\n📋 STEP 1: Loading Portfolio")
            if not self.load_portfolio(portfolio_file):
                return False
            
            # Step 2: Check data availability
            print("\n📁 STEP 2: Checking Cached Data")
            if not self.check_data_availability():
                print("\n⚠️  Missing cached data for some symbols")
                print("💡 To update data, run: python3 update_portfolio_data.py")
                return False
            
            # Step 3: Run analysis
            print("\n🔬 STEP 3: Running Analysis")
            if not self.run_analysis():
                return False
            
            # Step 4: Generate insights
            print("\n💡 STEP 4: Generating Insights")
            insights = self.generate_insights()
            if insights:
                print("✅ Insights generated successfully")
            else:
                print("⚠️  No insights generated")
            
            # Step 4b: Run Minervini Stage Analysis + Scoring + Signals
            print("\n🎯 STEP 4b: Minervini Stage Analysis & Composite Scoring")
            try:
                if hasattr(self, 'comprehensive_dataset') and not self.comprehensive_dataset.empty:
                    # Minervini stage classification (MUST run before scoring — scorer uses Stage/TT_Score)
                    minervini = MinerviniAnalyzer()
                    self.comprehensive_dataset = minervini.analyze_dataset(self.comprehensive_dataset)
                    
                    scorer = StockScorer()
                    self.comprehensive_dataset = scorer.score_dataset(self.comprehensive_dataset)
                    
                    engine = SignalEngine()
                    self.comprehensive_dataset = engine.generate_signals(self.comprehensive_dataset)
                    
                    print(f"✅ Scored {len(self.comprehensive_dataset)} stocks (avg composite: {self.comprehensive_dataset['Composite_Score'].mean():.1f})")
                    
                    # Re-save CSV/Excel with scoring columns
                    timestamp_str = date.today().strftime('%Y%m%d')
                    csv_path = f'reports/comprehensive_dataset_{timestamp_str}.csv'
                    self.comprehensive_dataset.to_csv(csv_path, index=False)
                    print(f"✅ Updated CSV with scoring columns: {csv_path}")
                    
                    # Re-init interactive filter with scored data
                    self.interactive_filter = InteractiveFilter(self.comprehensive_dataset)
            except Exception as e:
                print(f"⚠️  Scoring/signals: {e}")
            
            # Step 5: Create visualizations
            print("\n🎨 STEP 5: Creating Visualizations")
            charts = self.create_visualizations()
            if charts:
                print(f"✅ Created {len(charts)} charts")
            else:
                print("⚠️  No charts created")
            
            # Step 6: Generate HTML report
            print("\n📄 STEP 6: Generating Reports")
            html_report = self.generate_html_report()
            if html_report:
                print("✅ HTML report generated successfully")
            else:
                print("⚠️  HTML report generation failed")
            
            # Release chart objects after HTML report is written
            self.charts = {}
            gc.collect()
            
            # Step 7: Generate filtered reports (parallel if configured)
            print("\n🔍 STEP 7: Generating Filtered Reports")
            try:
                _filt = InteractiveFilter(self.comprehensive_dataset)
                filter_names = list(_filt.filter_criteria.keys())
                print(f"🔍 Generating {len(filter_names)} filtered reports...")
                filtered_reports = self.generate_multiple_filtered_reports(filter_names)
                if filtered_reports:
                    print(f"✅ Generated {len(filtered_reports)} filtered reports")
                else:
                    print("⚠️  No filtered reports generated")
            except Exception as e:
                print(f"❌ Error generating filtered reports: {e}")
            
            # Release filter objects before Step 8 heavy analysis
            self.interactive_filter = None
            gc.collect()
            
            # Step 8: Generate Advanced Analysis Reports
            print("\n🚀 STEP 8: Generating Advanced Analysis Reports")
            try:
                # Portfolio Drag Analysis
                print("\n🎯 Generating Portfolio Drag Analysis...")
                from pf_drag_analyzer import PortfolioDragAnalyzer
                drag_analyzer = PortfolioDragAnalyzer(self.comprehensive_dataset, self.historical_data)
                drag_report = drag_analyzer.generate_report()
                if drag_report:
                    print(f"✅ Portfolio drag analysis generated: {drag_report}")
                del drag_analyzer
                gc.collect()
                
                # Portfolio Optimization Report
                print("\n🔬 Generating Portfolio Optimization Report...")
                from pf_optimizer import PortfolioOptimizer
                optimizer = PortfolioOptimizer(self.comprehensive_dataset, self.historical_data)
                opt_report = optimizer.generate_report()
                if opt_report:
                    print(f"✅ Portfolio optimization report generated: {opt_report}")
                del optimizer
                gc.collect()
                
                # Minervini Stage Analysis Report
                print("\n📊 Generating Minervini Stage Analysis Report...")
                minervini = MinerviniAnalyzer()
                minervini_report = minervini.generate_report(self.comprehensive_dataset)
                if minervini_report:
                    print(f"✅ Minervini stage analysis generated: {minervini_report}")
                
                # Performance Bar Chart Report
                print("\n📊 Generating Performance Bar Chart Report...")
                from performance_bar_report import PerformanceBarReport
                perf_bar = PerformanceBarReport(self.comprehensive_dataset)
                perf_bar_report = perf_bar.generate_report()
                if perf_bar_report:
                    print(f"✅ Performance bar chart generated: {perf_bar_report}")
                
                # Documentation Report
                print("\n📚 Generating Documentation Report...")
                from report_documentation import ReportDocumentationGenerator
                doc_generator = ReportDocumentationGenerator()
                doc_report = doc_generator.generate_documentation()
                if doc_report:
                    print(f"✅ Documentation report generated: {doc_report}")
                    
            except Exception as e:
                print(f"❌ Error generating advanced reports: {e}")
                import traceback
                traceback.print_exc()
            
            # Step 8b: Generate Analytics Reports (Health, Alerts, Performance Trend)
            print("\n🧠 STEP 8b: Generating Analytics Reports")
            try:
                if hasattr(self, 'comprehensive_dataset') and not self.comprehensive_dataset.empty:
                    # Portfolio Health Dashboard
                    from portfolio_health import PortfolioHealthDashboard
                    health_dash = PortfolioHealthDashboard(self.comprehensive_dataset)
                    health_path = health_dash.generate_report()
                    if health_path:
                        print(f"✅ Health dashboard: {health_path}")
                    
                    # Alert Conditions Report
                    from alert_engine import AlertEngine
                    alert_eng = AlertEngine(self.comprehensive_dataset)
                    alert_path = alert_eng.generate_report()
                    if alert_path:
                        summary = alert_eng.get_summary()
                        print(f"✅ Alerts report: {alert_path} ({summary.get('total_alerts', 0)} alerts)")
                    
                    # Performance Trend Tracker
                    from performance_tracker import PerformanceTracker
                    tracker = PerformanceTracker(self.comprehensive_dataset)
                    tracker.save_run_metrics()
                    trend_path = tracker.generate_report()
                    if trend_path:
                        print(f"✅ Performance trend: {trend_path}")
            except Exception as e:
                print(f"⚠️  Analytics reports: {e}")
                import traceback
                traceback.print_exc()
            
            # Step 9: Generate Master Report (must be last — it discovers all other reports)
            print("\n📋 STEP 9: Generating Master Report")
            try:
                from master_report_generator import MasterReportGenerator
                master_generator = MasterReportGenerator()
                master_report_path = master_generator.save_master_report()
                print(f"✅ Master report generated: {master_report_path}")
            except Exception as e:
                print(f"❌ Error generating master report: {e}")
            
            print("\n" + "=" * 80)
            print("🎉 COMPLETE ANALYSIS WORKFLOW FINISHED!")
            print("=" * 80)
            
            # Display summary
            if self.insights and 'portfolio_summary' in self.insights:
                ps = self.insights['portfolio_summary']
                print(f"\n📊 PORTFOLIO SUMMARY:")
                print(f"   Total Stocks: {ps['total_stocks']}")
                print(f"   Portfolio Value: ₹{ps['current_value']:,.2f}")
                print(f"   Total Return: {ps['portfolio_return_pct']:.2f}%")
                print(f"   P&L: ₹{ps['total_pnl']:,.2f}")
            
            if hasattr(self, 'comprehensive_dataset') and 'Composite_Score' in self.comprehensive_dataset.columns:
                avg_score = self.comprehensive_dataset['Composite_Score'].mean()
                print(f"   Avg Composite Score: {avg_score:.1f}/100")
            
            print(f"\n📁 Files Generated:")
            timestamp = date.today().strftime('%Y%m%d')
            print(f"   📄 reports/index.html (🌟 Start here!)")
            print(f"   📄 reports/portfolio_report_{timestamp}.html")
            print(f"   📄 reports/portfolio_health_{timestamp}.html")
            print(f"   📄 reports/alert_conditions_{timestamp}.html")
            print(f"   📄 reports/performance_trend_{timestamp}.html")
            
            return success
            
        except Exception as e:
            print(f"\n❌ Error in complete analysis workflow: {e}")
            return False
    
    def _cleanup_old_reports(self):
        """Remove reports with date stamps different from today. Keeps 1 day only."""
        from config_manager import get_config
        cfg = get_config()
        retention_days = cfg.get_setting('report_settings.retention_days', 1)
        reports_dir = cfg.get_setting('system_settings.reports_directory', 'reports')
        today_stamp = date.today().strftime('%Y%m%d')
        
        if not os.path.exists(reports_dir):
            return
        
        removed = 0
        for f in os.listdir(reports_dir):
            fpath = os.path.join(reports_dir, f)
            if not os.path.isfile(fpath):
                continue
            # Check if file has a date stamp different from today
            # Pattern: any file containing _YYYYMMDD where date != today
            import re
            match = re.search(r'_(\d{8})\.', f)
            if match:
                file_date = match.group(1)
                if file_date != today_stamp:
                    try:
                        os.remove(fpath)
                        removed += 1
                    except OSError:
                        pass
        
        if removed > 0:
            print(f"🧹 Cleaned up {removed} old report files (keeping today: {today_stamp})")
        else:
            print(f"✅ No old reports to clean (retention: {retention_days} day)")
    
    def get_cache_info(self):
        """Display cache information"""
        try:
            cache_info = self.data_manager.get_cache_info()
            print(f"\n📁 Cache Information:")
            print(f"   Directory: {cache_info['cache_dir']}")
            print(f"   Total Files: {cache_info['total_files']}")
            
            if cache_info['files']:
                print(f"   Files:")
                for file_info in cache_info['files'][:5]:  # Show first 5 files
                    print(f"     - {file_info['name']} ({file_info['size_kb']} KB, {file_info['modified']})")
                
                if len(cache_info['files']) > 5:
                    print(f"     ... and {len(cache_info['files']) - 5} more files")
                
                return cache_info
            
        except Exception as e:
            print(f"❌ Error getting cache info: {e}")
            return {}
    
    def check_data_freshness_before_operation(self, operation_name: str = "operation") -> bool:
        """Check data freshness before any major operation"""
        try:
            from data_freshness_checker import DataFreshnessChecker
            
            print(f"🔍 Checking data freshness for {operation_name}...")
            checker = DataFreshnessChecker()
            results = checker.check_comprehensive_freshness()
            
            if results.get('update_needed', True):
                print(f"⚠️ Data update recommended before {operation_name}")
                if results.get('full_update_needed'):
                    print("🔄 Critical update needed - please run: python3 update_portfolio_data.py --force")
                    return False
                else:
                    print("⚡ Quick refresh recommended - please run: python3 update_portfolio_data.py")
                    return False
            else:
                print(f"✅ Data is fresh for {operation_name}")
                return True
                
        except Exception as e:
            print(f"⚠️ Error checking data freshness: {e}")
            return True  # Continue anyway
    
    def _get_current_portfolio_symbols(self) -> set:
        """Get the current set of stock symbols from the xlsx file."""
        if hasattr(self, 'portfolio_data') and self.portfolio_data is not None and not self.portfolio_data.empty:
            return set(self.portfolio_data['Symbol'].tolist())
        if hasattr(self, 'portfolio_manager') and hasattr(self.portfolio_manager, 'portfolio_data') and not self.portfolio_manager.portfolio_data.empty:
            return set(self.portfolio_manager.portfolio_data['Symbol'].tolist())
        return set()

    def generate_comprehensive_dataset(self, force_regenerate: bool = False) -> pd.DataFrame:
        """Generate comprehensive dataset with all technical indicators - SMART CACHE REUSE"""
        from datetime import date
        
        # Get authoritative stock list from xlsx (loaded in Step 1)
        xlsx_symbols = self._get_current_portfolio_symbols()
        
        # Check if comprehensive dataset already exists in memory
        if not force_regenerate and hasattr(self, 'comprehensive_dataset') and not self.comprehensive_dataset.empty:
            cached_symbols = set(self.comprehensive_dataset['Symbol'].tolist())
            if cached_symbols == xlsx_symbols:
                print("📊 Using existing in-memory comprehensive dataset...")
                return self.comprehensive_dataset
            else:
                added = xlsx_symbols - cached_symbols
                removed = cached_symbols - xlsx_symbols
                print(f"📋 Portfolio changed vs in-memory cache — regenerating dataset")
                if added:
                    print(f"   ➕ New stocks: {', '.join(sorted(added))}")
                if removed:
                    print(f"   ➖ Removed stocks: {', '.join(sorted(removed))}")
        
        # Check if comprehensive dataset exists in file cache
        timestamp = date.today().strftime('%Y%m%d')
        csv_filename = f'reports/comprehensive_dataset_{timestamp}.csv'
        
        if not force_regenerate and os.path.exists(csv_filename):
            try:
                print("🔍 Loading comprehensive dataset from file cache...")
                cached_df = pd.read_csv(csv_filename)
                cached_symbols = set(cached_df['Symbol'].tolist())
                
                # Validate symbols match the current xlsx
                if xlsx_symbols and cached_symbols != xlsx_symbols:
                    added = xlsx_symbols - cached_symbols
                    removed = cached_symbols - xlsx_symbols
                    print(f"⚠️  CSV cache stock list doesn't match xlsx — regenerating")
                    if added:
                        print(f"   ➕ New stocks ({len(added)}): {', '.join(sorted(added))}")
                    if removed:
                        print(f"   ➖ Removed stocks ({len(removed)}): {', '.join(sorted(removed))}")
                else:
                    self.comprehensive_dataset = cached_df
                    print(f"✅ SMART REUSE: Loaded {len(self.comprehensive_dataset)} stocks from cache")
                    return self.comprehensive_dataset
            except Exception as e:
                print(f"⚠️ Cache load failed: {e}, regenerating...")
        
        print("📊 Generating comprehensive dataset with advanced indicators...")
        
        try:
            # Ensure we have necessary data
            if not hasattr(self, 'portfolio_data') or self.portfolio_data is None:
                self.load_portfolio()
                self.portfolio_data = self.portfolio_manager.portfolio_data
            
            # Load historical data from cache
            if not self._load_cached_historical_data():
                print("❌ Failed to load cached data for dataset generation")
                print("💡 Run: python3 update_portfolio_data.py")
                return pd.DataFrame()
            
            # Generate comprehensive dataset
            dataset = self.technical_analyzer.generate_comprehensive_dataset(
                self.portfolio_data, 
                self.historical_data
            )
            
            # Ensure reports directory exists
            os.makedirs('reports', exist_ok=True)
            
            # Save as CSV
            dataset.to_csv(csv_filename, index=False)
            print(f"✅ CSV dataset saved: {csv_filename}")
            
            # Save as Excel with formatting
            excel_filename = f'reports/comprehensive_dataset_{timestamp}.xlsx'
            with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
                dataset.to_excel(writer, sheet_name='Portfolio_Dataset', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Portfolio_Dataset']
                
                # Format headers
                from openpyxl.styles import Font, PatternFill, Alignment
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
                
                for col_num, column in enumerate(dataset.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 20)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"✅ Excel dataset saved: {excel_filename}")
            
            # Store as instance attribute for other methods to use
            self.comprehensive_dataset = dataset
            
            # Initialize interactive filter system
            self.interactive_filter = InteractiveFilter(dataset)
            print("🔍 Interactive filter system initialized")
            
            # Display summary
            print(f"\n📋 Dataset Summary:")
            print(f"   📊 Total Stocks: {len(dataset)}")
            print(f"   📈 Total Indicators: {len(dataset.columns)}")
            print(f"   📄 Saved to: reports/comprehensive_dataset_{timestamp}.[csv/xlsx]")
            
            return dataset
            
        except Exception as e:
            print(f"❌ Error generating comprehensive dataset: {e}")
            return pd.DataFrame()
    
    def generate_filtered_report(self, filter_name: str = "All Stocks") -> str:
        """Generate filtered dataset report with interactive visualizations"""
        print(f"\n🔍 Generating filtered report: {filter_name}")
        
        try:
            # Ensure we have comprehensive dataset and filter system
            if self.comprehensive_dataset.empty:
                print("📊 Generating comprehensive dataset first...")
                self.generate_comprehensive_dataset()
            
            if self.interactive_filter is None:
                self.interactive_filter = InteractiveFilter(self.comprehensive_dataset)
            
            # OPTIMIZED: Use cached historical data (no duplicate fetch)
            if not self._load_cached_historical_data():
                print("❌ Failed to load historical data for filtered report")
                return ""
            
            # Generate filtered report
            report_path = self.interactive_filter.save_filtered_report(
                filter_name=filter_name,
                historical_data=self.historical_data
            )
            
            if report_path:
                print(f"✅ Filtered report generated: {report_path}")
            return report_path
            
        except Exception as e:
            print(f"❌ Error generating filtered report: {e}")
            return ""
    
    def list_available_filters(self) -> List[str]:
        """Get list of available filter criteria"""
        if self.interactive_filter is None:
            if self.comprehensive_dataset.empty:
                self.generate_comprehensive_dataset()
            self.interactive_filter = InteractiveFilter(self.comprehensive_dataset)
        
        return self.interactive_filter.get_available_filters()
    
    def show_filter_preview(self, filter_name: str) -> pd.DataFrame:
        """Preview filtered dataset without generating full report"""
        if self.interactive_filter is None:
            if self.comprehensive_dataset.empty:
                self.generate_comprehensive_dataset()
            self.interactive_filter = InteractiveFilter(self.comprehensive_dataset)
        
        return self.interactive_filter.apply_filter(filter_name)
    
    def generate_multiple_filtered_reports(self, filter_names: List[str] = None) -> List[str]:
        """Generate multiple filtered reports — parallel when configured"""
        if filter_names is None:
            essential_filters = [
                "Stocks Below WEMA30",
                "Stocks Above WEMA30", 
                "Near 52-Week High (within 5%)",
                "Oversold (RSI < 30)",
                "Overbought (RSI > 70)",
                "In Profit",
                "In Loss",
                "Stocks Below WEMA21",
                "Stocks Above WEMA21",
                "Stocks Below DSMA50",
                "Stocks Above DSMA50"
            ]
            try:
                if hasattr(self, 'comprehensive_dataset') and not self.comprehensive_dataset.empty:
                    filter_system = InteractiveFilter(self.comprehensive_dataset)
                    all_filter_names = list(filter_system.filter_criteria.keys())
                    filter_names = essential_filters + [f for f in all_filter_names if f not in essential_filters]
                else:
                    filter_names = essential_filters
            except Exception as e:
                print(f"⚠️  Using essential filters only: {e}")
                filter_names = essential_filters
        
        print(f"\n🔍 Generating {len(filter_names)} filtered reports...")
        
        # Remove stale filtered reports for today (will be regenerated; zero-result ones won't)
        from datetime import date as _date
        _today = _date.today().strftime('%Y%m%d')
        import glob as _glob
        for old in _glob.glob(f'reports/filtered_report_*_{_today}.html'):
            try:
                os.remove(old)
            except OSError:
                pass
        
        # --- Parallel report generation ---
        from config_manager import get_config
        cfg = get_config()
        use_parallel = cfg.get_setting('performance_settings.parallel_report_generation', True)
        max_workers = cfg.get_setting('performance_settings.max_workers', 4)
        
        # Limit parallelism to avoid OOM on low-memory machines
        # Each parallel worker creates its own InteractiveFilter copy (~1-2MB dataset)
        # plus historical_data reference + Plotly figures (~5-10MB per worker)
        if use_parallel and max_workers > 2:
            max_workers = min(max_workers, 2)
        
        report_paths = []
        successful_count = 0
        
        if use_parallel and len(filter_names) > 1:
            print(f"⚡ Parallel mode: {max_workers} workers")
            
            # Use reference for historical data (read-only), copy only the dataset
            _df_snapshot = self.comprehensive_dataset.copy()
            _hd_ref = self.historical_data if hasattr(self, 'historical_data') else None
            
            def _gen_one(fn):
                try:
                    # Each thread gets its own InteractiveFilter on a snapshot
                    from interactive_filter import InteractiveFilter as _IF
                    filt = _IF(_df_snapshot)
                    return filt.save_filtered_report(fn, _hd_ref, 'reports')
                except Exception as e:
                    import traceback
                    print(f"   ❌ [parallel] {fn}: {e}")
                    traceback.print_exc()
                    return None
            
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_name = {executor.submit(_gen_one, fn): fn for fn in filter_names}
                for future in as_completed(future_to_name):
                    fn = future_to_name[future]
                    try:
                        path = future.result()
                        if path:
                            report_paths.append(path)
                            successful_count += 1
                    except Exception as e:
                        print(f"   ❌ [parallel-future] {fn}: {e}")
        else:
            for i, filter_name in enumerate(filter_names):
                try:
                    print(f"📊 [{i+1}/{len(filter_names)}] Generating: {filter_name}")
                    path = self.generate_filtered_report(filter_name)
                    if path:
                        report_paths.append(path)
                        successful_count += 1
                        print(f"   ✅ {filter_name}")
                    else:
                        print(f"   ⚠️  {filter_name} - No data")
                except KeyboardInterrupt:
                    print(f"\n⚠️  Generation interrupted at {i+1}/{len(filter_names)}")
                    break
                except Exception as e:
                    print(f"   ❌ {filter_name}: {str(e)[:100]}...")
                    continue
        
        print(f"✅ Generated {successful_count} filtered reports successfully")
        return report_paths
    
    def generate_master_report(self):
        """Generate master report using MasterReportGenerator"""
        from master_report_generator import MasterReportGenerator
        
        print("🎯 Generating Master Report...")
        generator = MasterReportGenerator()
        master_report_path = generator.save_master_report()
        print(f"✅ Master report generated: {master_report_path}")
        return master_report_path

# Convenience functions
def generate_comprehensive_dataset_standalone() -> pd.DataFrame:
    """Standalone function to generate comprehensive dataset"""
    app = create_app()
    return app.generate_comprehensive_dataset()

def generate_filtered_reports(portfolio_file: str = None, filter_names: List[str] = None) -> List[str]:
    """Generate multiple filtered reports"""
    app = create_app(portfolio_file) if portfolio_file else create_app()
    return app.generate_multiple_filtered_reports(filter_names)

def show_available_filters(portfolio_file: str = None) -> List[str]:
    """Show all available filter criteria"""
    app = create_app(portfolio_file) if portfolio_file else create_app()
    filters = app.list_available_filters()
    
    print("🔍 Available Filter Criteria:")
    print("=" * 50)
    for i, filter_name in enumerate(filters, 1):
        print(f"   {i:2d}. {filter_name}")
    print("=" * 50)
    
    return filters

def create_filtered_report(filter_name: str, portfolio_file: str = None) -> str:
    """Create a single filtered report"""
    app = create_app(portfolio_file) if portfolio_file else create_app()
    return app.generate_filtered_report(filter_name)

# Convenience functions
def create_app(portfolio_file: str = None) -> PortfolioApp:
    """Create and return a new PortfolioApp instance"""
    if portfolio_file is None:
        portfolio_file = _default_portfolio_file()
    return PortfolioApp(portfolio_file)

def quick_analysis(portfolio_file: str = None) -> PortfolioApp:
    """Run a quick complete analysis using cached data and return the app instance
    
    Note: This uses cached data only. To update data first, run:
          python3 update_portfolio_data.py
    """
    app = create_app(portfolio_file) if portfolio_file else create_app()
    app.run_complete_analysis()
    return app

# Main execution
if __name__ == "__main__":
    print("🌟 Welcome to Portfolio Analysis System!")
    print("🎯 Lightweight Modular Portfolio Analysis with HTML Reporting")
    print("📊 Analysis Engine - Uses Cached Data Only")
    print("=" * 80)
    print()
    print("💡 To update market data first, run: python3 update_portfolio_data.py")
    print("=" * 80)
    
    try:
        # Get portfolio file from config
        portfolio_file = _default_portfolio_file()
        
        # Create app instance
        app = create_app()
        
        # Check if portfolio file exists
        if os.path.exists(portfolio_file):
            print(f"\n📁 Found portfolio file: {portfolio_file}")
            
            # Run complete analysis
            print("\n🚀 Running complete analysis...")
            success = app.run_complete_analysis()
            
            if success:
                print("\n✨ Analysis completed successfully!")
                print("\n💡 Next steps:")
                print("   - Open the generated HTML report in your browser")
                print("   - Review the portfolio summary and recommendations")
                print("   - Check individual stock analysis and signals")
                print("\n📅 To refresh with latest data:")
                print("   1. Run: python3 update_portfolio_data.py")
                print("   2. Run: python3 main_pf_app.py")
            else:
                print("\n⚠️  Analysis completed with some issues")
                print("\n💡 If data is missing, run: python3 update_portfolio_data.py")
                
            # Show cache info
            app.get_cache_info()
            
        else:
            print(f"\n⚠️  Portfolio file not found: {portfolio_file}")
            print("💡 Please ensure your portfolio Excel file is available.")
            print(f"🔧 Configure via config.json → system_settings.portfolio_file")
            print("   Or use: app = create_app('your_file.xlsx')")
            
    except KeyboardInterrupt:
        print("\n\n👋 Analysis interrupted by user. Goodbye!")
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
        print("💡 Please check your portfolio file format and try again.")
    
    print("\n🎯 Portfolio Analysis System - Complete!")
    print("=" * 80)
